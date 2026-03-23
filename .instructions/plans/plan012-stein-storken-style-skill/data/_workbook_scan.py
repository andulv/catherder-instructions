
import json, os
from pathlib import Path
from openpyxl import load_workbook

plan011 = Path(r"/workspace/catherder-instructions/.instructions/plans/plan011-svein-email-and-attachment-extraction")
attachments_root = plan011/'output/attachments'
with open(plan011/'output/manifests/candidate-documents.jsonl','r',encoding='utf-8') as f:
    records=[json.loads(line) for line in f if line.strip()]

xlsx=[]
for r in records:
    fn=(r.get('attachment_filename') or '').lower()
    sp=r.get('saved_path') or ''
    if fn.endswith('.xlsx') or sp.lower().endswith('.xlsx'):
        xlsx.append(r)

def rgb_of(fill):
    if not fill:
        return None
    color = fill.fgColor
    if color is None:
        return None
    if color.type == 'rgb' and color.rgb:
        return color.rgb
    if color.type == 'indexed':
        return f'indexed:{color.indexed}'
    if color.type == 'theme':
        return f'theme:{color.theme}'
    return None

out=[]
for r in xlsx:
    saved = r.get('saved_path')
    if not saved:
        continue
    p = Path(saved)
    if not p.exists():
        continue
    try:
        wb = load_workbook(p, data_only=False)
    except Exception as e:
        out.append({'saved_path': saved, 'attachment_filename': r.get('attachment_filename'), 'load_error': str(e)})
        continue
    item={
        'attachment_filename': r.get('attachment_filename'),
        'saved_path': saved,
        'confidence': r.get('confidence'),
        'candidate_class': r.get('candidate_class'),
        'evidence_codes': r.get('evidence_codes'),
        'sheet_count': len(wb.sheetnames),
        'sheet_names': wb.sheetnames,
        'first_sheet': None,
        'workbook_has_three_option_signal': False,
        'workbook_labels_sample': [],
        'format_signals': {},
    }
    color_counts={}
    bold=0
    merged=0
    freeze=[]
    labels=[]
    option_hits=0
    for ws in wb.worksheets:
        if ws.freeze_panes:
            freeze.append({'sheet': ws.title, 'freeze_panes': str(ws.freeze_panes)})
        merged += len(ws.merged_cells.ranges)
        rows_scanned=0
        nonempty=0
        text_cells=[]
        for row in ws.iter_rows():
            rows_scanned += 1
            if rows_scanned > 120:
                break
            for cell in row[:40]:
                v = cell.value
                if v is None or v == '':
                    continue
                nonempty += 1
                if isinstance(v, str):
                    sv=' '.join(v.split())
                    if sv and len(labels) < 25 and len(sv) <= 120:
                        labels.append({'sheet': ws.title, 'cell': cell.coordinate, 'text': sv})
                    low=sv.lower()
                    if 'alternativ' in low or low.strip() in {'a','b','c','1','2','3'}:
                        option_hits += 1
                    if len(sv) > 40:
                        text_cells.append({'cell': cell.coordinate, 'text': sv[:180]})
                if cell.font and cell.font.bold:
                    bold += 1
                rgb = rgb_of(cell.fill)
                if rgb:
                    color_counts[rgb] = color_counts.get(rgb,0)+1
        if item['first_sheet'] is None:
            item['first_sheet']={
                'title': ws.title,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'freeze_panes': str(ws.freeze_panes) if ws.freeze_panes else None,
                'merged_ranges': len(ws.merged_cells.ranges),
                'sample_text_cells': text_cells[:8],
                'appears_summary_like': any(any(k in (t['text'].lower()) for k in ['oppsummer', 'summary', 'alternativ', 'kommentar', 'forutset', 'anbefal']) for t in text_cells[:12]) or ws.max_column <= 12,
                'contains_explanatory_text': len(text_cells) >= 2,
            }
    top_colors = sorted(color_counts.items(), key=lambda kv: kv[1], reverse=True)[:8]
    item['workbook_has_three_option_signal'] = option_hits >= 3
    item['workbook_labels_sample'] = labels[:20]
    item['format_signals'] = {
        'top_fill_colors': [{'color': c, 'count': n} for c,n in top_colors],
        'bold_cell_count_scanned': bold,
        'merged_range_count': merged,
        'freeze_panes': freeze,
    }
    out.append(item)
print(json.dumps(out, ensure_ascii=False, indent=2))
